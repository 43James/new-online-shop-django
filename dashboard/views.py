from decimal import Decimal
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.http import Http404, HttpResponseRedirect, JsonResponse
from django.core.paginator import Paginator
from app_linebot.views import notify_admin_order_status, notify_user_approved, notify_user_pay_confirmed, send_acknowledge_notification, send_restock_notification
from django.db.models import Q, Sum, Max, F, ExpressionWrapper, DecimalField
from shop.models import Category, MonthlyStockRecord, Product, Stock, Subcategory, Suppliers, Total_Quantity, Receiving
from accounts.models import MyUser, Profile, WorkGroup
from orders.models import Order, Issuing, OutOfStockNotification
from .forms import AddProductForm, AddCategoryForm, AddSubcategoryForm, ApprovePayForm, EditCategoryForm, EditProductForm, ApproveForm, AddSuppliersForm, EditSubcategoryForm, EditSuppliersForm, EditWorkGroupForm, MonthYearForm, OrderFilterForm,  ReceivingForm, RecordMonthlyStockForm, WorkGroupForm
from django.db.models import F
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from datetime import datetime, timedelta
import openpyxl, re
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Border, Side, Alignment
from django.http import HttpResponse
from collections import defaultdict
import calendar
import locale
import plotly.express as px
import pandas as pd
locale.setlocale(locale.LC_TIME, 'th_TH.UTF-8')  # ตั้งค่า locale เป็นภาษาไทย
from .forms import UploadFileForm
from django.http import Http404
from openpyxl import Workbook
from babel.dates import format_datetime
from django.utils.dateparse import parse_date
from django.db.models import Subquery, OuterRef
from django.utils import timezone



def thai_month_name(month):
    thai_months = [
        'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน',
        'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 'สิงหาคม',
        'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม'
    ]
    return thai_months[month - 1] if 1 <= month <= 12 else ''

def convert_to_buddhist_era(year):
    return year + 543


def count_pending_orders():
    # ดึงข้อมูลออเดอร์ทั้งหมดที่รอการยืนยัน
    pending_orders = Order.objects.filter(status=None)
    return pending_orders.count()


    
from django.core.exceptions import PermissionDenied

def is_manager(user):
    if not user.is_manager:
        raise PermissionDenied
    return True

def is_executive(user):
    if not user.is_executive:
        raise PermissionDenied
    return True

def is_admin(user):
    if not user.is_admin:
        raise PermissionDenied
    return True

def is_authorized(user):
    # ถ้าผู้ใช้เป็น is_manager, is_executive หรือ is_admin อย่างใดอย่างหนึ่ง
    if user.is_manager or user.is_warehouse_manager or user.is_executive or user.is_admin:
        return True
    raise PermissionDenied

def is_authorized_manager(user):
    if user.is_manager or user.is_warehouse_manager or user.is_admin:
        return True
    raise PermissionDenied


# นำออกที่สมบูรณ์1
# รายงาน
@user_passes_test(is_authorized)
@login_required
def export_to_excel(request, month=None, year=None):
    now = datetime.now()

    # อ่านค่าจาก query string
    if request.GET.get('month'):
        month = int(request.GET.get('month'))
    if request.GET.get('year'):
        year = int(request.GET.get('year'))

    if not month:
        month = now.month
    if not year:
        year = now.year

    # กำหนดช่วงวันที่
    start_date = datetime(year, month, 1)
    end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(seconds=1)

    previous_month_end = start_date - timedelta(seconds=1)

    if timezone.is_naive(start_date):
        start_date = timezone.make_aware(start_date)
    if timezone.is_naive(end_date):
        end_date = timezone.make_aware(end_date)

    # category_id = request.GET.get('category')
    # if category_id:
    #     products = Product.objects.filter(category__category_id=category_id)
    #     selected_category_name = Category.objects.get(id=category_id).name_cate
    # else:
    #     products = Product.objects.all()
    #     selected_category_name = "ทั้งหมด"

    category_id = request.GET.get('category')
    selected_category_name = "ทั้งหมด"

    try:
        category_id = int(category_id)
    except (TypeError, ValueError):
        category_id = None

    if category_id:
        try:
            selected_category = Category.objects.get(id=category_id)
            selected_category_name = selected_category.name_cate
            products = Product.objects.filter(category__category_id=category_id)
        except Category.DoesNotExist:
            products = Product.objects.all()
            selected_category_name = "ทั้งหมด"
    else:
        products = Product.objects.all()

    report_data = []
    all_users = set()
    total_issued_value = Decimal(0)
    total_issued_value_by_user = defaultdict(Decimal)
    total_remaining_value_sum = Decimal(0)  # ✅ เพิ่มตรงนี้

    for product in products:
        previous_balance_record = MonthlyStockRecord.objects.filter(
            product=product,
            month=previous_month_end.month,
            year=previous_month_end.year
        ).first()
        previous_balance = previous_balance_record.end_of_month_balance if previous_balance_record else 0

        next_balance_record = MonthlyStockRecord.objects.filter(
            product=product,
            month=end_date.month,
            year=end_date.year
        ).first()
        total_remaining_value = next_balance_record.total_price if next_balance_record else 0

        received_current_month = Receiving.objects.filter(
            product=product, date_received__range=(start_date, end_date)
        ).aggregate(total_received=Sum('quantityreceived'))['total_received'] or 0

        total_balance = previous_balance + received_current_month

        issued_current_month = Issuing.objects.filter(
            product=product, datecreated__range=(start_date, end_date), order__status=True
        ).aggregate(total_issued=Sum('quantity'))['total_issued'] or 0

        remaining_balance = total_balance - issued_current_month

        total_remaining_value2 = Receiving.objects.filter(
            product=product
        ).aggregate(
            total_remaining=Sum(F('quantity') * F('unitprice'), output_field=DecimalField())
        )['total_remaining'] or Decimal(0)

        # ✅ สะสมยอดรวมมูลค่าวัสดุคงเหลือ
        if total_remaining_value:
            total_remaining_value_sum += total_remaining_value
        else:
            total_remaining_value_sum += total_remaining_value2

        issued_items = Issuing.objects.filter(
            product=product, datecreated__range=(start_date, end_date), order__status=True
        )
        total_issued_value_product = issued_items.aggregate(
            total_cost=Sum(F('price') * F('quantity'))
        )['total_cost'] or 0

        user_issuings = defaultdict(int)
        for issuing in issued_items:
            user_full_name = issuing.order.user.get_first_name()
            user_issuings[user_full_name] += issuing.quantity
            all_users.add(user_full_name)
            total_issued_value_by_user[user_full_name] += issuing.price * issuing.quantity
            total_issued_value += issuing.price * issuing.quantity

        report_data.append({
            'product_id': product.product_id,
            'product': product.product_name,
            'unit': product.unit,
            'previous_balance': previous_balance,
            'received_current_month': received_current_month,
            'total_balance': total_balance,
            'user_issuings': user_issuings,
            'issued_current_month': issued_current_month,
            'remaining_balance': remaining_balance,
            'total_issued_value': total_issued_value_product,
            'total_remaining_value': total_remaining_value,
            'total_remaining_value2': total_remaining_value2,
            'note': '',
        })

    sorted_users = sorted(all_users)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    font_title = Font(name='TH SarabunPSK', size=16, bold=True)
    font_header = Font(name='TH SarabunPSK', size=14, bold=True)
    font_data = Font(name='TH SarabunPSK', size=12)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    thai_month = thai_month_name(month)
    buddhist_year = convert_to_buddhist_era(year)

    # Header rows
    header_rows = [
        [f'รายการเบิกวัสดุ ({selected_category_name})'],
        [f'ประจำเดือน {thai_month} พ.ศ. {buddhist_year} ประจำปีงบประมาณ พ.ศ. {buddhist_year}'],
        ['หน่วยงาน โครงการอุทยานวิทยาศาสตร์ มหาวิทยาลัยอุบลราชธานี']
    ]
    for idx, row in enumerate(header_rows, start=1):
        worksheet.append(row)
        worksheet.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=16)
        cell = worksheet.cell(row=idx, column=1)
        cell.font = font_title
        cell.alignment = align_center

    # Column headers
    headers = ['รหัสวัสดุ', 'รายการสินค้า', 'หน่วยนับ', 'จำนวนคงเหลือ (ยกมา)', 'จำนวนรับเข้า (ปัจจุบัน)', 'รวมจำนวนคงเหลือบวกรับเข้า']
    headers += sorted_users
    headers += ['รวมจำนวนที่เบิก', 'มูลค่าที่เบิก (บาท)', 'จำนวนวัสดุคงเหลือ', 'มูลค่าวัสดุคงเหลือ (บาท)', 'หมายเหตุ']
    worksheet.append(headers)

    for cell in worksheet[4]:
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border

    # Data rows
    for item in report_data:
        user_values = [f"{item['user_issuings'].get(user, 0):,}" for user in sorted_users]
        remaining_value = f"{item['total_remaining_value2']:,}" if item['total_remaining_value2'] else f"{item['total_remaining_value']:,}"
        row = [
            item['product_id'],
            item['product'],
            item['unit'],
            f"{item['previous_balance']:,}",
            f"{item['received_current_month']:,}",
            f"{item['total_balance']:,}",
            *user_values,
            f"{item['issued_current_month']:,}",
            f"{item['total_issued_value']:,}",
            f"{item['remaining_balance']:,}",
            remaining_value,
            item['note']
        ]
        worksheet.append(row)

    # ✅ แถว: รวมรายจ่ายที่เบิก
    total_row_num = worksheet.max_row + 1
    total_row = [''] * 6 + [f'{total_issued_value_by_user[user]:,.2f} บาท' for user in sorted_users] + [''] + [f'{total_issued_value:,.2f} บาท'] + [''] + ['']
    worksheet.append(total_row)
    worksheet.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=6)
    cell = worksheet.cell(row=total_row_num, column=1, value="รวมรายจ่ายที่เบิก")
    cell.font = font_header
    cell.alignment = align_center

    # ✅ แถว: รวมมูลค่าวัสดุคงเหลือทั้งหมด
    total_remaining_row_num = worksheet.max_row + 1
    remaining_total_row = [''] * (7 + len(sorted_users)) + [''] + [''] + [f'{total_remaining_value_sum:,.2f} บาท'] + ['']
    worksheet.append(remaining_total_row)
    worksheet.merge_cells(start_row=total_remaining_row_num, start_column=1, end_row=total_remaining_row_num, end_column=7)
    cell = worksheet.cell(row=total_remaining_row_num, column=1, value="รวมมูลค่าวัสดุคงเหลือทั้งหมด")
    cell.font = font_header
    cell.alignment = align_center

    # Styling
    for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = font_data
            cell.alignment = align_center
            cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=monthly_report_{month}_{year}.xlsx'
    workbook.save(response)
    return response


@user_passes_test(is_authorized)
@login_required
# รายงานประจำเดือน
def monthly_report(request, month=None, year=None):
    now = datetime.now()

    if request.GET.get('month'):
        month = int(request.GET.get('month'))
    if request.GET.get('year'):
        year = int(request.GET.get('year'))

    if not month:
        month = now.month
    if not year:
        year = now.year

    month_name = thai_month_name(month)
    buddhist_year = convert_to_buddhist_era(year)

    # กำหนดวันที่เริ่มต้นและสิ้นสุด
    start_date = datetime(year, month, 1)
    end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(seconds=1)

    # ตรวจสอบว่า start_date เป็น naive datetime ก่อนทำให้มันมี timezone
    if timezone.is_naive(start_date):
        start_date = timezone.make_aware(start_date)

    # ตรวจสอบว่า end_date เป็น naive datetime ก่อนทำให้มันมี timezone
    if timezone.is_naive(end_date):
        end_date = timezone.make_aware(end_date)
    
    previous_month_start = (start_date - timedelta(days=1)).replace(day=1)
    previous_month_end = start_date - timedelta(seconds=1)

    category_id = request.GET.get('category')
    if category_id:
        products = Product.objects.filter(category__category_id=category_id)
        selected_category_name = Category.objects.get(id=category_id).name_cate
    else:
        products = Product.objects.all()
        selected_category_name = "ทั้งหมด"

    report_data = []
    all_users = set()
    total_issued_value = Decimal(0)
    total_issued_value_by_user = defaultdict(Decimal)
    total_remaining_value_sum = Decimal(0)  # ✅ รวมมูลค่าวัสดุคงเหลือทั้งหมด

    total_issued_quantity = Issuing.objects.filter(
        datecreated__range=(start_date, end_date),
        order__status=True
    ).aggregate(total_issued=Sum('quantity'))['total_issued'] or 0

    for product in products:
        previous_balance_record = MonthlyStockRecord.objects.filter(
            product=product,
            month=previous_month_end.month,
            year=previous_month_end.year
        ).first()

        previous_balance = previous_balance_record.end_of_month_balance if previous_balance_record else 0

        # ดึงข้อมูลยอดคงเหลือของเดือนถัดไป
        next_balance_record = MonthlyStockRecord.objects.filter(
            product=product,
            month=end_date.month,  # เดือนถัดไป
            year=end_date.year     # ปีของเดือนถัดไป
        ).first()

        # ใช้ข้อมูล total_price ของเดือนถัดไป หากไม่พบข้อมูลให้ใช้ค่า 0
        total_remaining_value = next_balance_record.total_price if next_balance_record else 0

        received_current_month = Receiving.objects.filter(
            product=product, date_received__range=(start_date, end_date)
        ).aggregate(total_received=Sum('quantityreceived'))['total_received'] or 0

        total_balance = previous_balance + received_current_month

        issued_current_month = Issuing.objects.filter(
            product=product, datecreated__range=(start_date, end_date), order__status=True
        ).aggregate(total_issued=Sum('quantity'))['total_issued'] or 0

        remaining_balance = total_balance - issued_current_month

        # คำนวณมูลค่าวัสดุคงเหลือ (total_remaining_value)
        total_remaining_value2 = Receiving.objects.filter(
            product=product
        ).aggregate(
            total_remaining=Sum(F('quantity') * F('unitprice'), output_field=DecimalField())
        )['total_remaining'] or Decimal(0)
        # print(f"Product ID: {product.product_id}, Total Remaining Value: {total_remaining_value2}")

        if total_remaining_value:
            total_remaining_value_sum += total_remaining_value
        else:
            total_remaining_value_sum += total_remaining_value2

        issued_items = Issuing.objects.filter(
            product=product, datecreated__range=(start_date, end_date), order__status=True
        )
        total_issued_value_product = issued_items.aggregate(total_cost=Sum(F('price') * F('quantity')))['total_cost'] or 0

        user_issuings = defaultdict(int)

        
        for issuing in issued_items:
            user_full_name = issuing.order.user.get_first_name()
            user_issuings[user_full_name] += issuing.quantity
            all_users.add(user_full_name)

            total_cost = issuing.price * issuing.quantity
            total_issued_value_by_user[user_full_name] += total_cost
            total_issued_value += total_cost
        

        report_data.append({
            'product_id': product.product_id,
            'product': product.product_name,
            'unit': product.unit,
            'previous_balance': previous_balance,
            'received_current_month': received_current_month,
            'total_balance': total_balance,
            'user_issuings': user_issuings,
            'issued_current_month': issued_current_month,
            'remaining_balance': remaining_balance,
            'total_issued_value': total_issued_value_product,
            'total_remaining_value': total_remaining_value,
            'total_remaining_value2': total_remaining_value2,
            'note': '',
        })

    sorted_all_users = sorted(all_users)

    context = {
        'title': 'รายงานประจำเดือน',
        'report_data': report_data,
        'all_users': sorted_all_users,
        'month': month,
        'year': year,
        'now': now,
        'years': range(2023, now.year + 1),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
        ],
        # 'pending_orders_count': count_pending_orders(),
        'month_name': month_name,
        'buddhist_year': buddhist_year,
        'total_issued_value_by_user': total_issued_value_by_user,
        'total_issued_value': total_issued_value,
        'total_issued_quantity': total_issued_quantity,
        'selected_category': category_id,
        'selected_category_name': selected_category_name,
        'total_remaining_value_sum': total_remaining_value_sum,  # ✅ ส่งไป template
    }
    return render(request, 'monthly_report.html', context)



@user_passes_test(is_authorized)
@login_required
def record_monthly_stock_view(request):
    if request.method == 'POST':
        form = RecordMonthlyStockForm(request.POST)
        if form.is_valid() and form.cleaned_data['confirm']:
            now = datetime.now()
            first_day_of_current_month = now.replace(day=1)
            last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)

            # ตรวจสอบว่ามีการบันทึกข้อมูลสำหรับเดือนก่อนหน้าแล้วหรือไม่
            existing_records = MonthlyStockRecord.objects.filter(
                month=last_day_of_previous_month.month,
                year=last_day_of_previous_month.year
            )

            if existing_records.exists():
                messages.error(request, 'มีการบันทึกข้อมูลของแล้ว.')
                return redirect('dashboard:record_monthly_stock')

            # สร้างบันทึกข้อมูลสำหรับเดือนก่อนหน้า
            products = Product.objects.all()
            for product in products:
                # คำนวณ total_quantity_received และ total_price จาก Receiving ที่เกี่ยวข้อง
                receiving_data = Receiving.objects.filter(
                    product=product,
                    # month=last_day_of_previous_month.month,
                    # year=last_day_of_previous_month.year
                ).aggregate(
                    total_quantity_received=Sum('quantity'),
                    total_remaining_value=Sum(F('quantity') * F('unitprice'), output_field=DecimalField())
                )

                total_quantity_received = receiving_data['total_quantity_received'] or 0
                total_price = receiving_data['total_remaining_value'] or 0.00

                MonthlyStockRecord.objects.create(
                    product=product,
                    month=last_day_of_previous_month.month,
                    year=last_day_of_previous_month.year,
                    end_of_month_balance=total_quantity_received,
                    total_price=total_price
                )

            messages.success(request, 'บันทึกข้อมูลสต็อกสินค้าเรียบร้อยแล้ว.')
            return redirect('dashboard:monthly_stock_records')
    else:
        form = RecordMonthlyStockForm()

    context = {
        'title': 'บันทึกยอดวัสดุคงเหลือประจำเดือน',
        'form': form,
        'pending_orders_count': count_pending_orders(),
    }

    return render(request, 'record_monthly_stock.html', context)



@user_passes_test(is_authorized)
@login_required
# รายงรายงานจำนวนคงเหลือ ยกมา
def monthly_stock_records(request):
    now = datetime.now()

    # Determine the default month and year, which is the previous month
    if now.month == 1:
        default_month = 12
        default_year_ad = now.year - 1
    else:
        default_month = now.month - 1
        default_year_ad = now.year

    # Get month and year from GET parameters, defaulting to the previous month
    month = int(request.GET.get('month', default_month))
    year_buddhist = int(request.GET.get('year', default_year_ad + 543))

    # Convert Buddhist year to A.D. for database queries
    year_ad = year_buddhist - 543

    # Filter MonthlyStockRecord by the selected month and year, ordered by product ID
    records = MonthlyStockRecord.objects.filter(month=month, year=year_ad).order_by('product__id')

    # Prepare the context dictionary
    context = {
        'title': 'ข้อมูลวัสดุคงเหลือ (ยกมา)',
        'records': records,
        'selected_month': month,
        'selected_year': year_buddhist,
        'years': range(2023 + 543, datetime.now().year + 1 + 543),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
        ],
        'pending_orders_count': count_pending_orders(),
    }
    # เพิ่มโค้ดส่วนนี้เข้าไปในฟังก์ชัน monthly_stock_records ที่อยู่ใน views.py
    previous_month = month if month > 1 else 12
    previous_year_ad = year_ad if month > 1 else year_ad - 1
    context['previous_month_name'] = thai_month_name(previous_month)
    context['previous_year_buddhist'] = convert_to_buddhist_era(previous_year_ad)
    return render(request, 'monthly_stock_records.html', context)


from dateutil.relativedelta import relativedelta
@user_passes_test(is_authorized)
@login_required
# รายงงานจำนวนวัสดุคงเหลือ ตามไตรมาส
def monthly_stock_sum(request):
    now = datetime.now()

    # การกรองตามหมวดหมู่หลัก
    selected_category_id = request.GET.get('category')
    selected_category = Category.objects.filter(id=selected_category_id).first() if selected_category_id else None

    # ใช้เดือนและปีของเดือนที่แล้วเป็นค่าเริ่มต้น
    last_month = now - relativedelta(months=1)
    month = int(request.GET.get('start_month', last_month.month))
    year_buddhist = int(request.GET.get('start_year', last_month.year + 543))
    end_month = int(request.GET.get('end_month', now.month))
    end_year_buddhist = int(request.GET.get('end_year', now.year + 543))
    quarter = int(request.GET.get('quarter', 0))

    year_ad = year_buddhist - 543
    end_year_ad = end_year_buddhist - 543

    quarter_text = ""
    month_text = ""

    # กำหนดการแสดงผลปีตามเงื่อนไขที่เลือก
    if quarter != 0:
        if quarter == 1:
            month, end_month = 10, 12
            year_ad = end_year_ad = now.year - 1  # ตั้งค่าให้ปีเป็นปีที่แล้ว (ต.ค.-ธ.ค.)
            year_buddhist = end_year_buddhist = now.year - 1 + 543
            quarter_text = "ไตรมาสที่ 1"
            month_text = "ตุลาคม - ธันวาคม"
        elif quarter == 2:
            month, end_month = 1, 3
            quarter_text = "ไตรมาสที่ 2"
            month_text = "มกราคม - มีนาคม"
        elif quarter == 3:
            month, end_month = 4, 6
            quarter_text = "ไตรมาสที่ 3"
            month_text = "เมษายน - มิถุนายน"
        elif quarter == 4:
            month, end_month = 7, 9
            quarter_text = "ไตรมาสที่ 4"
            month_text = "กรกฎาคม - กันยายน"
        year_range_text = f"{year_buddhist}"
    else:
        year_range_text = f"{year_buddhist} - {end_year_buddhist}"

    records = MonthlyStockRecord.objects.filter(month=month, year=year_ad).order_by('product__id')

    if selected_category:
        records = records.filter(product__category__category_id=selected_category.id)

    # เพิ่มการหา name_cate ของหมวดหมู่หลัก
    selected_category_name = Category.objects.get(id=selected_category_id).name_cate if selected_category_id else "ทั้งหมด"

    product_sums = {}
    for record in records:
        product_id = record.product.id
        if product_id not in product_sums:
            product_sums[product_id] = {
                'product_id': record.product.product_id,
                'product_name': record.product.product_name,
                'total_balance_quarter': 0,
                'total_price_quarter': 0,
                'unit': record.product.unit,
                'month': record.month,
                'year': record.year,
            }
        product_sums[product_id]['total_balance_quarter'] += record.end_of_month_balance
        product_sums[product_id]['total_price_quarter'] += record.total_price

    sorted_product_sums = dict(sorted(product_sums.items()))

    total_balance = sum([item['total_balance_quarter'] for item in sorted_product_sums.values()])
    total_price = sum([item['total_price_quarter'] for item in sorted_product_sums.values()])

     # คำนวณเดือนและปีที่แล้ว (เดือนก่อนหน้า) จากค่าที่ผู้ใช้เลือก
    if month == 1:
        previous_month = 12
        previous_year = year_ad - 1
    else:
        previous_month = month - 1
        previous_year = year_ad

    context = {
        'title': 'วัสดุคงเหลือประจำเดือน',
        'product_sums': sorted_product_sums,
        'selected_start_month': month,
        'selected_start_year': year_buddhist,
        'selected_end_month': end_month,
        'selected_end_year': end_year_buddhist,
        'quarter': quarter,
        'quarter_text': quarter_text,
        'month_text': month_text,
        'year_range_text': year_range_text,            
        'selected_category': selected_category_id,
        'selected_categorys': selected_category_name,
        'years': range(2023 + 543, datetime.now().year + 1 + 543),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
        ],
        'total_balance': total_balance,
        'total_price': total_price,
        'pending_orders_count': count_pending_orders(),
    }

    # คำนวณเดือนและปีที่แล้ว (เดือนก่อนหน้า)
    previous_month = month if month > 1 else 12
    previous_year = year_ad if month > 1 else year_ad - 1
    context['previous_month_name'] = thai_month_name(previous_month)  # ฟังก์ชันที่แปลงเดือนเป็นชื่อเดือนภาษาไทย
    context['previous_year_buddhist'] = convert_to_buddhist_era(previous_year)  # ฟังก์ชันที่แปลงปีคริสต์ศักราชเป็นปีพุทธศักราช
    return render(request, 'monthly_stock_sum.html', context)



def export_monthly_stock_sum_to_excel(request):
    now = datetime.now()
    last_month = now.month - 1 if now.month > 1 else 12
    last_year = now.year if now.month > 1 else now.year - 1

    month = int(request.GET.get('start_month', last_month))
    year_buddhist = int(request.GET.get('start_year', last_year + 543))
    end_month = int(request.GET.get('end_month', last_month))
    end_year_buddhist = int(request.GET.get('end_year', last_year + 543))
    quarter = int(request.GET.get('quarter', 0))

    year_ad = year_buddhist - 543
    end_year_ad = end_year_buddhist - 543

    quarter_text = ""
    month_text = ""
    # year_range_text = f"{year_buddhist} - {end_year_buddhist}"

    # กำหนดการแสดงผลปีตามเงื่อนไขที่เลือก
    if quarter != 0:
        if quarter == 1:
            month, end_month = 10, 12
            year_ad = end_year_ad = now.year - 1  # ตั้งค่าให้ปีเป็นปีที่แล้ว (ต.ค.-ธ.ค.)
            year_buddhist = end_year_buddhist = now.year - 1 + 543
            quarter_text = "ไตรมาสที่ 1"
            month_text = "ตุลาคม - ธันวาคม"
        elif quarter == 2:
            month, end_month = 1, 3
            quarter_text = "ไตรมาสที่ 2"
            month_text = "มกราคม - มีนาคม"
        elif quarter == 3:
            month, end_month = 4, 6
            quarter_text = "ไตรมาสที่ 3"
            month_text = "เมษายน - มิถุนายน"
        elif quarter == 4:
            month, end_month = 7, 9
            quarter_text = "ไตรมาสที่ 4"
            month_text = "กรกฎาคม - กันยายน"
        year_range_text = f"{year_buddhist}"
    else:
        year_range_text = f"{year_buddhist} - {end_year_buddhist}"

    # การกรองตามหมวดหมู่หลัก
    selected_category_id = request.GET.get('category')
    selected_category = "ทั้งหมด"

    if selected_category_id and selected_category_id.isdigit():
        category = Category.objects.filter(id=int(selected_category_id)).first()
        if category:
            selected_category = category.name_cate  # ใช้ชื่อหมวดหมู่แทน

    records = MonthlyStockRecord.objects.filter(
        Q(month__gte=month) & Q(year__gte=year_ad) &
        Q(month__lte=end_month) & Q(year__lte=end_year_ad)
    )

    if selected_category_id and selected_category_id.isdigit() and category:
        records = records.filter(product__category__category_id=category.id)

    records = records.order_by('year', 'month')

    product_sums = {}
    for record in records:
        product_id = record.product.id
        if product_id not in product_sums:
            product_sums[product_id] = {
                'product_name': record.product.product_name,
                'total_balance_quarter': 0,
                'total_price_quarter': 0,
                'unit': record.product.unit,
            }
        product_sums[product_id]['total_balance_quarter'] += record.end_of_month_balance
        product_sums[product_id]['total_price_quarter'] += record.total_price

    sorted_product_sums = dict(sorted(product_sums.items()))

    total_balance = sum([item['total_balance_quarter'] for item in sorted_product_sums.values()])
    total_price = sum([item['total_price_quarter'] for item in sorted_product_sums.values()])

    # สร้างไฟล์ Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Stock Summary"

    # ตั้งค่าฟอนต์เป็น TH SarabunPSK
    font_th_sarabun = Font(name="TH SarabunPSK", size=14)

    # เพิ่มหัวเรื่อง
    title_text = f"ข้อมูลวัสดุคงเหลือ {quarter_text} {month_text} พ.ศ. {year_range_text} ( { selected_category } )"
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.font = Font(name="TH SarabunPSK", size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # เพิ่มหัวคอลัมน์
    headers = ['ลำดับ', 'วัสดุ', 'รวมจำนวนคงเหลือ', 'หน่วย', 'รวมจำนวนเงิน(บาท)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = font_th_sarabun

    # กำหนดรูปแบบเส้นตาราง
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # เพิ่มข้อมูลวัสดุคงเหลือ
    for row_num, (product_id, data) in enumerate(sorted_product_sums.items(), 3):
        ws.cell(row=row_num, column=1, value=row_num-2).font = font_th_sarabun  # ลำดับ
        ws.cell(row=row_num, column=2, value=data['product_name']).font = font_th_sarabun  # วัสดุ
        ws.cell(row=row_num, column=3, value=f"{data['total_balance_quarter']:,}").font = font_th_sarabun  # รวมจำนวนคงเหลือ
        ws.cell(row=row_num, column=4, value=data['unit']).font = font_th_sarabun  # หน่วย
        ws.cell(row=row_num, column=5, value=f"{data['total_price_quarter']:,}").font = font_th_sarabun  # รวมจำนวนเงิน

    # เพิ่มผลรวมทั้งหมด
    total_row_num = len(sorted_product_sums) + 3

    # รวมคอลัมน์ 1 และ 2 เข้าด้วยกัน
    ws.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=2)

    # ตั้งค่าเซลล์ที่รวมให้มีข้อความ "รวมทั้งหมด" และจัดให้อยู่ตรงกลาง
    cell = ws.cell(row=total_row_num, column=1, value="รวมทั้งหมด")
    cell.font = Font(name="TH SarabunPSK", size=14, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # เพิ่มข้อมูลในคอลัมน์อื่นๆ และตั้งค่าให้เป็นตัวหนา
    ws.cell(row=total_row_num, column=3, value=f"{total_balance:,}").font = Font(name="TH SarabunPSK", size=14, bold=True)
    ws.cell(row=total_row_num, column=4, value="ชิ้น").font = Font(name="TH SarabunPSK", size=14, bold=True)
    ws.cell(row=total_row_num, column=5, value=f"{total_price:,} บาท").font = Font(name="TH SarabunPSK", size=14, bold=True)

    # เพิ่มเส้นตารางให้กับ cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

    # ตั้งค่าความกว้างของคอลัมน์
    column_dimensions = {
        'A': 5,
        'B': 40,
        'C': 15,
        'D': 10,
        'E': 15,
    }
    for col, width in column_dimensions.items():
        ws.column_dimensions[col].width = width

    # บันทึกไฟล์ Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=monthly_stock_sum_{now.strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response



@user_passes_test(is_authorized)
@login_required
# export excel รายงงานจำนวนคงเหลือ ยกมา
def export_monthly_stock_records_to_excel(request):
    now = datetime.now()
    last_month = now.month - 1 if now.month > 1 else 12
    last_year = now.year if now.month > 1 else now.year - 1

    # ดึงเดือนและปีจากพารามิเตอร์ GET หรือใช้ค่าเดือนและปีของเดือนที่แล้ว
    month = int(request.GET.get('month', last_month))
    year_buddhist = int(request.GET.get('year', last_year + 543))

    # แปลงปี พ.ศ. เป็น ค.ศ. สำหรับการค้นหาในฐานข้อมูล
    year_ad = year_buddhist - 543

    # กรองข้อมูล MonthlyStockRecord ตามเดือนและปี
    records = MonthlyStockRecord.objects.filter(month=month, year=year_ad)

    # สร้าง response สำหรับการบันทึกไฟล์ Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="monthly_stock_records_{month}_{year_buddhist}.xlsx"'

    # สร้างไฟล์ Excel workbook
    wb = Workbook()
    ws = wb.active

    # ตั้งชื่อ sheet
    ws.title = f"Monthly Stock Records {month}-{year_buddhist}"  # ใช้ (-) เป็นตัวคั่น

    # ฟอนต์ที่ใช้
    thai_font = Font(name='TH Sarabun New', bold=True, size=16)
    regular_font = Font(name='TH Sarabun New', size=14)

    # เพิ่มข้อมูลหัวเรื่อง
    previous_month_name = thai_month_name(month)
    previous_year_buddhist = convert_to_buddhist_era(year_ad)
    ws.merge_cells('A1:H1')
    cell = ws.cell(row=1, column=1)
    cell.value = f"ข้อมูลสินค้าคงเหลือประจำเดือน {previous_month_name} พ.ศ. {previous_year_buddhist}"
    cell.font = thai_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # กำหนด header
    headers = ['ลำดับ', 'วัสดุ', 'เดือน', 'ปี', 'จำนวนคงเหลือ', 'หน่วย', 'จำนวนเงิน(บาท)', 'วันที่บันทึก']
    ws.append(headers)
    for cell in ws[2]:
        cell.font = thai_font
        cell.alignment = Alignment(horizontal='center')

    # กำหนดรูปแบบเส้นตาราง
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # เพิ่มข้อมูลลงใน rows
    for idx, record in enumerate(records, start=1):
        date_str = format_datetime(record.date_recorded, "d MMMM y H:mm", locale='th')
        row = [
            idx,
            record.product.product_name,
            record.month,
            record.year + 543,  # แปลงปี ค.ศ. เป็น พ.ศ.
            record.end_of_month_balance,
            record.product.unit,
            record.total_price,
            date_str  # แปลงวันที่เป็น string ในรูปแบบที่ต้องการ
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = regular_font

    # เพิ่มเส้นตารางให้กับ cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border

    # บันทึก workbook ลงใน response
    wb.save(response)
    return response





@user_passes_test(is_authorized)
@login_required
# รายงานรับเข้าประจำเดือน
def monthly_report_receive(request):
    now = datetime.now()

    # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    month = int(request.GET.get('month', now.month))
    year_buddhist = int(request.GET.get('year', now.year + 543))

    # แปลงปี พ.ศ. เป็น ค.ศ. สำหรับการค้นหาในฐานข้อมูล
    year_ad = year_buddhist - 543

    # ดึงข้อมูลรับเข้าสินค้าที่มีเดือนและปีที่ระบุ พร้อมกับข้อมูลหมวดหมู่หลักและหมวดหมู่ย่อย
    receiving_data = Receiving.objects.filter(
        month=month,
        year=year_ad
    ).select_related('product__category__category').values(
        'product__product_name',
        'product__product_id',
        'product__unit',
        'product__category__name_sub',  # หมวดหมู่ย่อย
        'product__category__category__name_cate',  # หมวดหมู่หลัก
        'suppliers__supname',
        'quantityreceived',
        'unitprice',
        'date_received'
    ).annotate(
        total_quantity=Sum('quantityreceived'),
        total_price=Sum(F('quantityreceived') * F('unitprice'))
    )

    # การกรองตามหมวดหมู่หลัก
    category_id = request.GET.get('category')
    selected_category_name = "ทั้งหมด"
    if category_id:
        receiving_data = receiving_data.filter(product__category__category__id=category_id)
        selected_category_name = Category.objects.get(id=category_id).name_cate

    # ดึงหมวดหมู่หลักทั้งหมดสำหรับการแสดงใน dropdown
    categories = Category.objects.all()

    # คำนวณรวมจำนวนเงินทั้งหมด
    total_price_sum = receiving_data.aggregate(total_sum=Sum('total_price'))['total_sum'] or 0

    # กำหนดค่าให้กับตัวแปร context
    context = {
        'title': 'รายการรับเข้าวัสดุ',
        'receiving_data': receiving_data,
        'selected_month': month,
        'selected_year': year_buddhist,
        'selected_category': category_id,
        'selected_category_name': selected_category_name,
        'categories': categories,
        'years': range(2023 + 543, datetime.now().year + 1 + 543),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
        ],
        'month_name': thai_month_name(month),
        'total_price_sum': total_price_sum,  # เพิ่มผลรวมจำนวนเงินทั้งหมด
        'pending_orders_count': count_pending_orders(),
    }

    previous_month = month - 1 if month > 1 else 12
    previous_year = year_ad if month > 1 else year_ad - 1
    context['previous_month_name'] = thai_month_name(previous_month)
    context['previous_year_buddhist'] = convert_to_buddhist_era(previous_year)
    
    return render(request, 'monthly_report_receive.html', context)



@user_passes_test(is_authorized)
@login_required
# export excel รายงานรับเข้าประจำเดือน
def export_to_excel_receive(request):
    now = datetime.now()

    # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    last_month = now.month if now.month > 1 else 12
    last_year = now.year if now.month > 1 else now.year - 1

    # ตรวจสอบว่ามีการระบุเดือนและปีในพารามิเตอร์ GET หรือไม่ ถ้าไม่มีใช้เดือนและปีของเดือนที่แล้ว
    month = int(request.GET.get('month', last_month))
    year_buddhist = int(request.GET.get('year', last_year + 543))

    # แปลงปี พ.ศ. เป็น ค.ศ. สำหรับการค้นหาในฐานข้อมูล
    year_ad = year_buddhist - 543

    # ดึงข้อมูลรับเข้าสินค้าที่มีเดือนและปีที่ระบุ
    receiving_data = Receiving.objects.filter(
        month=month,
        year=year_ad
    ).select_related('product', 'suppliers').values(
        'product__product_name',
        'product__unit',
        'product__category__category__name_cate',  # หมวดหมู่หลัก
        'product__category__name_sub',  # หมวดหมู่ย่อย
        'suppliers__supname',
        'quantityreceived',
        'unitprice',
        'date_created'
    ).annotate(
        total_quantity=Sum('quantityreceived'),
        total_price=Sum(F('quantityreceived') * F('unitprice'))
    )

    # ตรวจสอบ category_id ว่ามีค่าและเป็นตัวเลขหรือไม่
    category_id = request.GET.get('category')
    selected_category_name = "ทั้งหมด"
    if category_id and category_id.isdigit():
        receiving_data = receiving_data.filter(product__category__category__id=int(category_id))
        selected_category_name = Category.objects.get(id=int(category_id)).name_cate

    # สร้าง response สำหรับการบันทึกไฟล์ Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="receiving_report_{month}_{year_buddhist}.xlsx"'

    # สร้างไฟล์ Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # ตั้งชื่อ sheet
    ws.title = f"Receiving Report {month}-{year_buddhist}"

    thai_font = Font(name='TH Sarabun New', bold=True, size=16)
    regular_font = Font(name='TH Sarabun New', size=14)

    # เพิ่มข้อมูลหัวเรื่อง
    previous_month_name = thai_month_name(month)
    previous_year_buddhist = convert_to_buddhist_era(year_ad)
    ws.merge_cells('A1:J1')
    cell = ws.cell(row=1, column=1)
    cell.value = f"รายการรับเข้าวัสดุ ประจำเดือน {previous_month_name} พ.ศ. {previous_year_buddhist} หมวดหมู่ที่เลือก : {selected_category_name}"
    cell.font = thai_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # กำหนด header
    headers = ['ลำดับ', 'รายการวัสดุ', 'หมวดหมู่หลัก', 'หมวดหมู่ย่อย', 'ซัพพลายเออร์', 'จำนวนที่รับเข้า', 'หน่วย', 'ราคา/หน่วย', 'รวมจำนวนเงิน', 'วันที่รับเข้า']
    ws.append(headers)
    for cell in ws[2]:
        cell.font = thai_font
        cell.alignment = Alignment(horizontal='center')

    # กำหนดรูปแบบเส้นตาราง
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # เพิ่มข้อมูลลงใน rows
    for idx, data in enumerate(receiving_data, start=1):
        date_created = data['date_created']
        date_str = format_datetime(date_created, "d MMMM y H:mm", locale='th') if date_created else "ไม่ระบุ"
        row = [
            idx,
            data['product__product_name'],
            data['product__category__category__name_cate'],  # หมวดหมู่หลัก
            data['product__category__name_sub'],  # หมวดหมู่ย่อย
            data['suppliers__supname'],
            "{:,}".format(data['quantityreceived']),
            data['product__unit'],
            "{:,.2f}".format(data['unitprice']),
            "{:,.2f}".format(data['total_price']),
            date_str
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = regular_font

    # เพิ่มเส้นตารางให้กับ cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = thin_border

    # บันทึก workbook ลงใน response
    wb.save(response)
    return response



@user_passes_test(is_authorized)
@login_required
# รายงานยอดรวมที่มีการเบิกแต่ละเดือน
def report_monthly_totals(request):
    # orders = Order.objects.all()
    # กรองเฉพาะ order ของผู้ใช้งานปัจจุบันที่เข้าสู่ระบบ และมีสถานะออเดอร์เป็น True
    orders = Order.objects.filter(status=True)
    report_monthly_totals = defaultdict(Decimal)
    
    for order in orders:
        for item in order.items.all():
            # ใช้ tuple (order.month, order.year) เป็น key
            report_monthly_totals[(order.month, order.year)] += item.get_cost()
    
    # แปลงค่าใน report_monthly_totals เป็น float เพื่อให้สามารถใช้งานใน template ได้
    report_monthly_totals_float = {(month, year): float(total) for (month, year), total in report_monthly_totals.items()}
    
    context = {
        'title': 'Monthly Totals',
        'report_monthly_totals': report_monthly_totals_float,
        'pending_orders_count': count_pending_orders(),
    }
    return render(request, 'report_monthly_totals.html', context)



@user_passes_test(is_authorized)
@login_required
def total_quantity(request):
    products = Product.objects.all()
    stock = Total_Quantity.objects.all()

    query = request.GET.get('q')
    if query is not None:
        lookups = Q(product__product_id__icontains=query) | Q(product__product_name__icontains=query)
        stock = Total_Quantity.objects.filter(lookups)

    page = request.GET.get('page')

    p = Paginator(stock, 20)
    try:
        stock = p.page(page)
    except:
        stock = p.page(1)

    context = {
        'title':'รวมจำนวนวัสดุที่รับเข้า', 
        'stock':stock,
        'product':products,
        'pending_orders_count': count_pending_orders(),
        }
    return render(request, 'total_quantity.html', context)


@user_passes_test(is_authorized)
@login_required
# หน้าออเดอร์แต่ละคนในเดือนนั่นๆในเอชบอร์ด
def report_order_list(request):
    now = datetime.now()

    # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    month = int(request.GET.get('month', now.month))
    year_buddhist = int(request.GET.get('year', now.year + 543))

    # แปลงปี พ.ศ. เป็น ค.ศ. สำหรับการค้นหาในฐานข้อมูล
    year_ad = year_buddhist - 543

    # ดึงข้อมูลรับเข้าสินค้าที่มีเดือนและปีที่ระบุ
    order_data = Order.objects.filter(
        month=month,
        year=year_ad
     ).select_related('user')

    # กำหนดค่าให้กับตัวแปร context
    context = {
        'title': 'รายการเบิกประจำเดือน รายบุคคล',
        'order_data': order_data,
        'selected_month': month,
        'selected_year': year_buddhist,
        'years': range(2023 + 543, datetime.now().year + 1 + 543),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
        ],
        'month_name': thai_month_name(month),
        'pending_orders_count': count_pending_orders(),
    }

    previous_month = month - 1 if month > 1 else 12
    previous_year = year_ad if month > 1 else year_ad - 1
    context['previous_month_name'] = thai_month_name(previous_month)
    context['previous_year_buddhist'] = convert_to_buddhist_era(previous_year)
    
    return render(request, 'report_order_list.html', context)


@user_passes_test(is_authorized)
@login_required
# export excel หน้าออเดอร์แต่ละคนในเดือนนั่นๆในเอชบอร์ด
def export_excel_order(request):
    month = int(request.GET.get('month'))
    year_buddhist = int(request.GET.get('year'))
    year_ad = year_buddhist - 543

    # ดึงข้อมูล Order ที่ตรงกับเดือนและปีที่ระบุ
    order_data = Order.objects.filter(
        month=month,
        year=year_ad
    ).select_related('user', 'user__profile')

    # สร้าง Workbook และ Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'รายการเบิกประจำเดือน รายบุคคล'

    thai_font = Font(name='TH Sarabun New', bold=True, size=16)
    regular_font = Font(name='TH Sarabun New', size=14)

    # เพิ่มข้อมูลหัวเรื่อง
    previous_month_name = thai_month_name(month)
    previous_year_buddhist = convert_to_buddhist_era(year_ad)
    ws.merge_cells('A1:H1')
    cell = ws.cell(row=1, column=1)
    cell.value = f"รายการเบิกประจำเดือน รายบุคคล ประจำเดือน {previous_month_name} พ.ศ. {previous_year_buddhist}"
    cell.font = thai_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # กำหนดหัวตาราง
    columns = ['ลำดับ', 'ชื่อผู้ใช้งาน', 'กลุ่มงาน', 'เลขที่เบิก', 'ยอดรวม', 'วันที่เบิก', 'สถานะ', 'หมายเหตุ']
    ws.append(columns)

    # กำหนดฟอนต์ภาษาไทย ขนาด 16
    font = Font(name='TH Sarabun New', size=14)
    alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col in ws.iter_cols(min_row=2, max_row=2, min_col=1, max_col=len(columns)):
        for cell in col:
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border

    # เพิ่มข้อมูลในตาราง
    for idx, order in enumerate(order_data, start=1):
        user_full_name = order.user.get_full_name()
        total_price = order.get_total_price
        date_created = format_datetime(order.date_created, "d MMMM y เวลา H:mm น.", locale='th') if order.date_created else "ไม่ระบุ"
        status = 'อนุมัติ' if order.status else 'ปฏิเสธ' if order.status is False else 'รอดำเนินการ..'
        other = order.other or ''
        workgroup = str(order.user.profile.workgroup)

        row = [idx, user_full_name, workgroup, order.id, total_price, date_created, status, other]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = regular_font

        # เพิ่มเส้นตารางให้กับ cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border

    # สร้าง HTTP Response และบันทึกไฟล์ Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_orders_users_{month}_{year_buddhist}.xlsx'
    wb.save(response)
    return response



@user_passes_test(is_authorized)
@login_required
# หน้าเดชบอร์ด
def dashboard_home(request):
    now = datetime.now()
    
    # การกรองตามหมวดหมู่หลัก
    selected_category_id = request.GET.get('category')
    selected_category = Category.objects.filter(id=selected_category_id).first() if selected_category_id else None

     # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    current_month = now.month
    current_year_buddhist = now.year + 543

    # ดึงค่าจาก GET หรือใช้ค่าปัจจุบัน
    start_month = int(request.GET.get('start_month', current_month))
    start_year_buddhist = int(request.GET.get('start_year', current_year_buddhist))
    end_month = int(request.GET.get('end_month', current_month))
    end_year_buddhist = int(request.GET.get('end_year', current_year_buddhist))

    # แปลงปี พ.ศ. เป็น ค.ศ.
    start_year_ad = start_year_buddhist - 543
    end_year_ad = end_year_buddhist - 543

    # ตรวจสอบค่าที่ได้รับ
    print(f"Start Month: {start_month}, Start Year AD: {start_year_ad}")
    print(f"End Month: {end_month}, End Year AD: {end_year_ad}")
    
    # แปลงปี พ.ศ. เป็น ค.ศ. สำหรับการค้นหาในฐานข้อมูล
    start_year_ad = start_year_buddhist - 543
    end_year_ad = end_year_buddhist - 543

    # ตรวจสอบค่าที่ได้รับ
    print(f"Start Month: {start_month}, Start Year AD: {start_year_ad}")
    print(f"End Month: {end_month}, End Year AD: {end_year_ad}")

    # กรองข้อมูลการเบิกวัสดุที่ได้รับการอนุมัติออเดอร์แล้ว ตามช่วงเวลาที่เลือก
    issuing_data = Issuing.objects.filter(order__status=True)

    if start_year_ad == end_year_ad:
        # กรณีปีเดียวกัน
        issuing_data = issuing_data.filter(
            year=start_year_ad,
            month__gte=start_month,
            month__lte=end_month
        )
    else:
        # กรณีข้ามปี
        issuing_data = issuing_data.filter(
            Q(year=start_year_ad, month__gte=start_month) |
            Q(year=end_year_ad, month__lte=end_month)
        )

    if selected_category:
        issuing_data = issuing_data.filter(product__category__category_id=selected_category.id)

    # เพิ่มการหา name_cate ของหมวดหมู่หลัก
    selected_category_name = Category.objects.get(id=selected_category_id).name_cate if selected_category_id else "ทั้งหมด"

    issuing_data = issuing_data.values(
        'order__user__profile__workgroup__work_group',
        'product__product_name',
        'product__unit'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('price') * F('quantity'))
    ).order_by('order__user__profile__workgroup__work_group')

    # ข้อมูลสำหรับ Context
    if issuing_data.exists():
        df = pd.DataFrame(issuing_data)

        # เตรียมข้อมูลสำหรับ Chart.js
        labels = df['order__user__profile__workgroup__work_group'].unique().tolist()
        products = df['product__product_name'].unique().tolist()
        
        # Data สำหรับกราฟปริมาณการเบิกวัสดุ
        datasets_quantity = []
        colors = [
            'rgba(75, 192, 192, 1)', 'rgba(255, 99, 132, 1)', 'rgba(54, 162, 235, 1)',
            'rgba(255, 206, 86, 1)', 'rgba(153, 102, 255, 1)', 'rgba(255, 159, 64, 1)'
        ]
        # สร้างแม็พสีสำหรับแต่ละผลิตภัณฑ์
        color_map = {product: get_random_color() for product in products}

        for i, product in enumerate(products):
            data = df[df['product__product_name'] == product]
            monthly_totals = [data[data['order__user__profile__workgroup__work_group'] == label]['total_quantity'].sum() for label in labels]
            color = color_map[product]
            datasets_quantity.append({
                'label': product,
                'data': monthly_totals,
                'borderColor': color,
                'backgroundColor': color.replace('1)', '0.8)'),
                'fill': False
            })

        # Data สำหรับกราฟจำนวนเงินที่เบิก
        datasets_amount = []
        for i, product in enumerate(products):
            data = df[df['product__product_name'] == product]
            monthly_totals = [data[data['order__user__profile__workgroup__work_group'] == label]['total_amount'].sum() for label in labels]
            color = color_map[product]
            datasets_amount.append({
                'label': product,
                'data': monthly_totals,
                'borderColor': color,
                'backgroundColor': color.replace('1)', '0.8)'),
                'fill': False
            })

        chart_data_quantity = {
            'labels': labels,
            'datasets': datasets_quantity
        }

        chart_data_amount = {
            'labels': labels,
            'datasets': datasets_amount
        }

        context = {
            'title': 'Dashboard',
            'chart_data_quantity': json.dumps(chart_data_quantity, ensure_ascii=False, default=str),
            'chart_data_amount': json.dumps(chart_data_amount, ensure_ascii=False, default=str),
            'pending_orders_count': count_pending_orders(),
            'now': now,
            'selected_start_month': start_month,
            'selected_start_year': start_year_buddhist,
            'selected_end_month': end_month,
            'selected_end_year': end_year_buddhist,
            'selected_category': selected_category_id,
            'selected_categorys': selected_category_name,
            'categories': Category.objects.all(),
            'years': range(2023 + 543, datetime.now().year + 1 + 543),
            'months': [
                (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
                (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
                (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
            ],
        }
    else:
        context = {
            'title': 'Dashboard',
            'chart_data_quantity': json.dumps({'labels': [], 'datasets': []}, ensure_ascii=False, default=str),
            'chart_data_amount': json.dumps({'labels': [], 'datasets': []}, ensure_ascii=False, default=str),
            'pending_orders_count': count_pending_orders(),
            'now': now,
            'selected_start_month': start_month,
            'selected_start_year': start_year_buddhist,
            'selected_end_month': end_month,
            'selected_end_year': end_year_buddhist,
            'selected_category': selected_category_id,
            'selected_categorys': selected_category_name,
            'categories': Category.objects.all(),
            'years': range(2023 + 543, datetime.now().year + 1 + 543),
            'months': [
                (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
                (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
                (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
            ],
        }

    return render(request, 'dashboard_home.html', context)


import random
def get_random_color():
    r = random.randint(0, 255)
    g = random.randint(0, 255)
    b = random.randint(0, 255)
    return f'rgba({r}, {g}, {b}, 0.5)'  # ปรับ alpha เป็น 0.2 เพื่อให้สีอ่อนลง

@user_passes_test(is_authorized)
@login_required
def dashboard_report(request):
    now = datetime.now()

    # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    current_month = now.month
    current_year_buddhist = now.year + 543

    start_month = int(request.GET.get('start_month', current_month))
    start_year_buddhist = int(request.GET.get('start_year', current_year_buddhist))
    end_month = int(request.GET.get('end_month', current_month))
    end_year_buddhist = int(request.GET.get('end_year', current_year_buddhist))
    selected_category_id = request.GET.get('category')  # รับค่าหมวดหมู่หลักจากพารามิเตอร์ GET

    start_year_ad = start_year_buddhist - 543
    end_year_ad = end_year_buddhist - 543

    # การกรองข้อมูลตามเดือนและปี
    if start_year_ad == end_year_ad:
        # กรณีที่ปีเริ่มต้นและปีสิ้นสุดเป็นปีเดียวกัน
        combined_data = Issuing.objects.filter(
            order__status=True,
            year=start_year_ad,
            month__gte=start_month,
            month__lte=end_month
        )
    else:
        # กรณีที่ปีเริ่มต้นและปีสิ้นสุดเป็นคนละปี
        combined_data = Issuing.objects.filter(
            order__status=True
        ).filter(
            Q(year=start_year_ad, month__gte=start_month) | 
            Q(year=end_year_ad, month__lte=end_month) |
            Q(year__gt=start_year_ad, year__lt=end_year_ad)
        )

    if selected_category_id:
        combined_data = combined_data.filter(product__category__id=selected_category_id)

    # เพิ่มการหา name_cate ของหมวดหมู่หลัก
    selected_category_name = Category.objects.get(id=selected_category_id).name_cate if selected_category_id else "ทั้งหมด"

    # การประมวลผลข้อมูลและการแปลงเป็น DataFrame
    combined_data = combined_data.values(
        'order__user__profile__workgroup__work_group',
        'product__product_name',
        'product__unit',
        'month',
        'year'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('price') * F('quantity'))
    ).order_by('order__user__profile__workgroup__work_group', 'month')

    # แปลงข้อมูลที่ได้มาเป็น DataFrame
    df = pd.DataFrame(list(combined_data))

    if not df.empty:
        df['month'] = df['month'].astype(int)
        df['total_amount'] = df['total_amount'].astype(float)
        df['total_quantity'] = df['total_quantity'].astype(float)

        # เปลี่ยนตัวเลขเป็นชื่อเดือน
        month_names = {
            1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน',
            5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฎาคม', 8: 'สิงหาคม',
            9: 'กันยายน', 10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'
        }

        labels = [month_names[m] for m in sorted(df['month'].unique().tolist())]
        work_groups = df['order__user__profile__workgroup__work_group'].unique().tolist()

        # สีสำหรับกลุ่มงานแต่ละกลุ่มงาน
        colors = [
            'rgba(75, 192, 192, 1)', 'rgba(255, 99, 132, 1)', 'rgba(54, 162, 235, 1)',
            'rgba(255, 206, 86, 1)', 'rgba(153, 102, 255, 1)', 'rgba(255, 159, 64, 1)'
        ]

        # กราฟแนวโน้มมูลค่าการเบิกวัสดุแต่ละกลุ่มงาน
        value_datasets = []
        for i, work_group in enumerate(work_groups):
            group_data = df[df['order__user__profile__workgroup__work_group'] == work_group]
            data = [group_data[group_data['month'] == month]['total_amount'].sum() for month in sorted(df['month'].unique())]
            value_datasets.append({
                'label': f'{work_group}',
                'data': data,
                'borderColor': colors[i % len(colors)],
                'backgroundColor': colors[i % len(colors)].replace('1)', '1)'),
                'fill': False
            })

        value_chart_data = {
            'labels': labels,
            'datasets': value_datasets
        }

        # กราฟแนวโน้มการใช้วัสดุแต่ละกลุ่มงานในแต่ละเดือน
        quantity_datasets = []
        for i, work_group in enumerate(work_groups):
            group_data = df[df['order__user__profile__workgroup__work_group'] == work_group]
            data = [group_data[group_data['month'] == month]['total_quantity'].sum() for month in sorted(df['month'].unique())]
            quantity_datasets.append({
                'label': f'{work_group}',
                'data': data,
                'borderColor': colors[i % len(colors)],
                'backgroundColor': colors[i % len(colors)].replace('1)', '1)'),
                'fill': False
            })

        quantity_chart_data = {
            'labels': labels,
            'datasets': quantity_datasets
        }

    else:
        value_chart_data = {
            'labels': [],
            'datasets': []
        }
        quantity_chart_data = {
            'labels': [],
            'datasets': []
        }

    context = {
        'title': 'รายงานแผนภูมิเส้น',
        'title2': 'กราฟแนวโน้มมูลค่าการเบิกวัสดุและการใช้วัสดุแต่ละกลุ่มงาน',
        'pending_orders_count': count_pending_orders(),
        'now': now,
        'selected_start_month': start_month,
        'selected_start_year': start_year_buddhist,
        'selected_end_month': end_month,
        'selected_end_year': end_year_buddhist,
        'selected_category': selected_category_id,  # เพิ่มการส่งข้อมูลหมวดหมู่หลัก
        'selected_categorys': selected_category_name,
        'years': range(2023 + 543, datetime.now().year + 1 + 543),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
        ],
        'value_chart_data': json.dumps(value_chart_data, ensure_ascii=False, default=str),
        'quantity_chart_data': json.dumps(quantity_chart_data, ensure_ascii=False, default=str),
        'categories': Category.objects.all(),  # เพิ่มข้อมูลหมวดหมู่หลัก
        'current_month': current_month,  # เพิ่มข้อมูลเดือนปัจจุบัน
        'current_year_buddhist': current_year_buddhist  # เพิ่มข้อมูลปีพ.ศ. ปัจจุบัน
    }
    return render(request, 'dashboard_report.html', context)





def issuing_report(request):
    now = datetime.now()

    # การกรองตามกลุ่มงาน
    selected_workgroup_id = request.GET.get('workgroup')
    selected_workgroup = WorkGroup.objects.filter(id=selected_workgroup_id).first() if selected_workgroup_id else None

     # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    current_month = now.month
    current_year_buddhist = now.year + 543

    # ดึงค่าจาก GET หรือใช้ค่าปัจจุบัน
    start_month = int(request.GET.get('start_month', current_month))
    start_year_buddhist = int(request.GET.get('start_year', current_year_buddhist))
    end_month = int(request.GET.get('end_month', current_month))
    end_year_buddhist = int(request.GET.get('end_year', current_year_buddhist))

    # แปลงปี พ.ศ. เป็น ค.ศ.
    start_year_ad = start_year_buddhist - 543
    end_year_ad = end_year_buddhist - 543

    # ตรวจสอบค่าที่ได้รับ
    print(f"Start Month: {start_month}, Start Year AD: {start_year_ad}")
    print(f"End Month: {end_month}, End Year AD: {end_year_ad}")


    # กรองข้อมูลการเบิกวัสดุที่ได้รับการอนุมัติออเดอร์แล้ว ตามช่วงเวลาที่เลือก
    issuing_data = Issuing.objects.filter(
        order__status=True
    )

    if start_year_ad == end_year_ad:
        # กรณีที่อยู่ในปีเดียวกัน
        issuing_data = issuing_data.filter(
            order__year=start_year_ad,
            order__month__gte=start_month,
            order__month__lte=end_month
        )
    else:
        # กรณีที่ข้ามปี
        issuing_data = issuing_data.filter(
            Q(order__year=start_year_ad, order__month__gte=start_month) |
            Q(order__year=end_year_ad, order__month__lte=end_month)
        )

    if selected_workgroup:
        issuing_data = issuing_data.filter(order__user__profile__workgroup__id=selected_workgroup.id)

    # เพิ่มการหา work_group ของกลุ่มงาน
    selected_workgroup_name = WorkGroup.objects.get(id=selected_workgroup_id).work_group if selected_workgroup_id else "ทั้งหมด"

    issuing_data = issuing_data.values(
        'order__user__profile__workgroup__work_group',
        'order__user__first_name',
        'product__product_name',
        'product__category__category__name_cate',
        'product__unit'
    ).annotate(
        total_quantity=Sum('quantity')
    ).order_by('order__user__profile__workgroup__work_group', 'order__user__first_name')

    # ข้อมูลสำหรับ Context
    if issuing_data.exists():
        df = pd.DataFrame(issuing_data)

        # เตรียมข้อมูลสำหรับ Chart.js
        labels = df['order__user__profile__workgroup__work_group'].unique().tolist()
        users = df['order__user__first_name'].unique().tolist()

        # Data สำหรับกราฟแท่ง
        datasets_quantity = []
        colors = [
            'rgba(75, 192, 192, 0.2)', 'rgba(255, 99, 132, 0.2)', 'rgba(54, 162, 235, 0.2)',
            'rgba(255, 206, 86, 0.2)', 'rgba(153, 102, 255, 0.2)', 'rgba(255, 159, 64, 0.2)'
        ]
        # color_map = {user: colors[i % len(colors)] for i, user in enumerate(users)}
        
        # สร้างแม็พสีสำหรับแต่ละผลิตภัณฑ์
        color_map = {user: get_random_color() for user in users}

        for user in users:
            data = df[df['order__user__first_name'] == user]
            quantities = [data[data['order__user__profile__workgroup__work_group'] == label]['total_quantity'].sum() for label in labels]
            color = color_map[user]
            datasets_quantity.append({
                'label': user,
                'data': quantities,
                'backgroundColor': color,
                'borderColor': color.replace('0.2', '1'),  # เปลี่ยน alpha เป็น 1 เพื่อให้ขอบสีชัดเจน
                'borderWidth': 1
            })

        chart_data_quantity = {
            'labels': labels,
            'datasets': datasets_quantity
        }

        context = {
            'title': 'รายงานการเบิกวัสดุ',
            'chart_data_quantity': json.dumps(chart_data_quantity, ensure_ascii=False, default=str),
            'now': now,
            'issuing_data': issuing_data,  # เพิ่มบรรทัดนี้
            'selected_start_month': start_month,
            'selected_start_year': start_year_buddhist,
            'selected_end_month': end_month,
            'selected_end_year': end_year_buddhist,
            'selected_workgroup': selected_workgroup_id,
            'selected_workgroup_name': selected_workgroup_name,
            'workgroups': WorkGroup.objects.all(),
            'pending_orders_count': count_pending_orders(),
            'years': range(2023 + 543, datetime.now().year + 1 + 543),
            'months': [
                (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
                (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
                (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
            ],
        }
    else:
        context = {
            'title': 'รายงานการเบิกวัสดุ',
            'chart_data_quantity': json.dumps({'labels': [], 'datasets': []}, ensure_ascii=False, default=str),
            'now': now,
            'issuing_data': issuing_data,  # เพิ่มบรรทัดนี้
            'selected_start_month': start_month,
            'selected_start_year': start_year_buddhist,
            'selected_end_month': end_month,
            'selected_end_year': end_year_buddhist,
            'selected_workgroup': selected_workgroup_id,
            'selected_workgroup_name': selected_workgroup_name,
            'workgroups': WorkGroup.objects.all(),
            'pending_orders_count': count_pending_orders(),
            'years': range(2023 + 543, datetime.now().year + 1 + 543),
            'months': [
                (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
                (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
                (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')
            ],
        }

    return render(request, 'issuing_report.html', context)



@user_passes_test(is_authorized)
@login_required
def export_products_to_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Define headers
    headers = ['หมวดหมู่หลัก', 'หมวดหมู่ย่อย', 'รหัสพัสดุ', 'ชื่อรายการ', 'คำอธิบาย', 'หน่วย', 
               'คงเหลือเดือนก่อน', 'สต๊อกคงเหลือเดือนนี้', 'วันที่เพิ่มรายการ']
    ws.append(headers)

    # Query products
    products = Product.objects.all().select_related('category__category')  # ใช้ prefetch_related สำหรับดึงข้อมูล category และ subcategory

    # Add data rows
    for product in products:
        # Get stock for the previous month
        last_month_record = MonthlyStockRecord.objects.filter(product=product, month=(timezone.now().month - 1), year=timezone.now().year).first()
        last_month_balance = last_month_record.end_of_month_balance if last_month_record else 0
        
        # Get stock for the current month
        current_month_record = MonthlyStockRecord.objects.filter(product=product, month=timezone.now().month, year=timezone.now().year).first()
        current_month_balance = current_month_record.end_of_month_balance if current_month_record else 0
        
        # Get the total quantity received for the product from the Receiving model
        total_quantity = Receiving.total_quantity_by_product(product.id)  # ใช้ฟังก์ชัน total_quantity_by_product

        # Add data row
        row = [
            product.category.category.name_cate if product.category and product.category.category else '',  # หมวดหมู่หลัก
            product.category.name_sub if product.category else '',  # หมวดหมู่ย่อย
            product.product_id,
            product.product_name,
            product.description,
            product.unit,
            last_month_balance,  # คงเหลือเดือนก่อน
            total_quantity,  # สต๊อกคงเหลือเดือนนี้
            product.date_created.strftime('%Y-%m-%d %H:%M:%S'),  # วันที่เพิ่มรายการ
        ]
        ws.append(row)

    # Save workbook to response
    wb.save(response)
    return response


@login_required
def products(request):
    query = request.GET.get('q')
    filter_stock = request.GET.get('filter_stock', 'all')
    products = Product.objects.all()

    if query is not None:
        lookups = Q(product_id__icontains=query) | Q(product_name__icontains=query)
        products = products.filter(lookups)
        
    # ดึง end_of_month_balance จาก MonthlyStockRecord ของเดือนล่าสุด
    # ⚠️ โค้ดนี้จะทำงานได้ต่อเมื่อคุณมีโมเดล MonthlyStockRecord ที่ถูกต้อง
    latest_balance_subquery = MonthlyStockRecord.objects.filter(
        product=OuterRef('pk')
    ).order_by('-year', '-month').values('end_of_month_balance')[:1]

    # ใช้ annotate เพื่อคำนวณสต็อกคงเหลือปัจจุบันและข้อมูลอื่น ๆ
    products = products.annotate(
        # ✅ คำนวณสต็อกคงเหลือปัจจุบัน (Current Stock) จาก Receiving.quantity
        total_current_stock=Sum('Receiving__quantity'), 
        
        # คำนวณอื่น ๆ
        # Note: ถ้า Receiving__quantity คือยอดคงเหลือ, total_quantity_received จะไม่ถูกต้อง 
        # ให้สมมติว่าโค้ดนี้ต้องการใช้ total_current_stock เป็นยอดคงเหลือหลัก
        total_quantity_received=Sum('Receiving__quantity'), # คงไว้ตามโค้ดต้นฉบับ
        latest_receiving_date=Max('Receiving__date_created'),
        total_remaining_value=Sum(F('Receiving__quantity') * F('Receiving__unitprice'), output_field=DecimalField()),
        end_of_month_balance=Subquery(latest_balance_subquery)
    )

    # กรองวัสดุตามตัวเลือกจากปุ่ม โดยใช้ total_current_stock
    if filter_stock == 'in_stock':
        # ✅ กรอง: สต็อกมากกว่า 0
        products = products.filter(total_current_stock__gt=0)
        # ✅ เรียง: ตามสต็อกคงเหลือ
        products = products.order_by('total_current_stock') 
    elif filter_stock == 'out_of_stock':
        # ✅ กรอง: สต็อกเท่ากับ 0 หรือ NULL (ใช้ Q object เพื่อรวม NULL เข้าไปด้วย)
        products = products.filter(Q(total_current_stock=0) | Q(total_current_stock__isnull=True))
    else:
        # ✅ เรียง: ตาม ID หรือ total_current_stock
        products = products.order_by('id') # เปลี่ยนจาก total_current_stock เพื่อให้แสดงรายการทั้งหมดตาม ID

    # เรียงลำดับสุดท้าย (หากต้องการ)
    # products = products.order_by('total_quantity_received')

    # pagination
    page = request.GET.get('page')
    p = Paginator(products, 20)
    try:
        products = p.page(page)
    except Exception:
        products = p.page(1)

    context = {
        'title': 'รายการวัสดุทั้งหมด',
        'products': products,
        'pending_orders_count': count_pending_orders(),
        # ✅ แก้ไข: ใช้ total_current_stock สำหรับการคำนวณยอดรวมที่แสดงในหน้า
        'total_quantity': sum(product.total_current_stock or 0 for product in products), 
        'filter_stock': filter_stock
    }
    return render(request, 'products.html', context)


@user_passes_test(is_authorized_manager)
@login_required
# def products(request):
#     query = request.GET.get('q')
#     filter_stock = request.GET.get('filter_stock', 'all')  # รับค่าจากปุ่มกรอง
#     products = Product.objects.all()

#     if query is not None:
#         lookups = Q(product_id__icontains=query) | Q(product_name__icontains=query)
#         products = products.filter(lookups)

#     # กรองวัสดุตามตัวเลือกจากปุ่ม
#     if filter_stock == 'in_stock':
#         products = products.filter(quantityinstock__gt=0)  # แสดงเฉพาะวัสดุที่ยังมีในสต๊อก
#         products = products.order_by('quantityinstock')  # เรียงจากน้อยไปมาก
#     elif filter_stock == 'out_of_stock':
#         products = products.filter(quantityinstock=0)  # แสดงเฉพาะวัสดุที่หมดสต๊อก
#     else:
#         products = products.order_by('id')  # เรียงตามสต็อกทั้งหมด

#      # ดึง end_of_month_balance จาก MonthlyStockRecord ของเดือนล่าสุด
#     latest_balance_subquery = MonthlyStockRecord.objects.filter(
#         product=OuterRef('pk')
#     ).order_by('-year', '-month').values('end_of_month_balance')[:1]

#     # ใช้ annotate เพื่อคำนวณจำนวนรวมและวันที่รับเข้าล่าสุด
#     products = products.annotate(
#         total_quantity_received=Sum('Receiving__quantity'),
#         latest_receiving_date=Max('Receiving__id'),
#         total_remaining_value=Sum(F('Receiving__quantity') * F('Receiving__unitprice'), output_field=DecimalField()),
#         end_of_month_balance=Subquery(latest_balance_subquery)  # ดึงจำนวนคงเหลือเดือนล่าสุด
#     )

#     # เพิ่มการเรียงลำดับที่แน่นอนหลังการ annotate
#     # products = products.order_by('id')
#     # เรียงลำดับจากน้อยไปมากตาม total_quantity_received
#     products = products.order_by('total_quantity_received')

#     # pagination
#     page = request.GET.get('page')
#     p = Paginator(products, 20)
#     try:
#         products = p.page(page)
#     except:
#         products = p.page(1)

#     context = {
#         'title': 'รายการวัสดุทั้งหมด',
#         'products': products,
#         'pending_orders_count': count_pending_orders(),
#         'total_quantity': sum(product.quantityinstock for product in products),
#         'filter_stock': filter_stock  # เก็บค่าตัวเลือกการกรองไว้ใน context
#     }
#     return render(request, 'products.html', context)


@user_passes_test(is_authorized_manager)
@login_required
def upload_products(request):
    if request.method == 'POST':
        file = request.FILES['file']
        try:
            df = pd.read_excel(file)
            for index, row in df.iterrows():
                main_category_name = row['หมวดหมู่หลัก']
                sub_category_name = row['หมวดหมู่ย่อย']

                # Create or get the main category
                main_category, created = Category.objects.get_or_create(name_cate=main_category_name)

                # Create or get the subcategory
                sub_category, created = Subcategory.objects.get_or_create(name_sub=sub_category_name, category=main_category)

                # Create the product
                Product.objects.create(
                    category=sub_category,
                    product_id=row['รหัสพัสดุ'],
                    product_name=row['ชื่อรายการ'],
                    description=row['คำอธิบาย'],
                    unit=row['หน่วย'],
                )
            messages.success(request, 'นำเข้าข้อมูลสำเร็จ')
        except Exception as e:
            messages.error(request, f'เกิดข้อผิดพลาด: {e}')
    return redirect('dashboard:products')


@user_passes_test(is_authorized_manager)
@login_required
def add_product(request):
    if request.method == 'POST':
        form = AddProductForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                form.save()
                messages.success(request, 'เพิ่มวัสดุเรียบร้อยแล้ว!')
                return redirect('dashboard:add_product')
        except Exception as e:
            messages.error(request, f'เกิดข้อผิดพลาด: {e}')
    else:
        form = AddProductForm()
    context = {'title':'เพิ่มวัสดุ', 
               'form':form, 
                'Category':Category.objects.all(),
                'Subcategory':Subcategory.objects.all(),
                'pending_orders_count': count_pending_orders(),}
    return render(request, 'add_product.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def delete_product(request, id):
    product = Product.objects.filter(id=id).delete()
    messages.success(request, 'ลบวัสดุสำเร็จ')
    return redirect('dashboard:products')



@user_passes_test(is_authorized_manager)
@login_required
def edit_product(request, id):
    product = get_object_or_404(Product, id=id)
    if request.method == 'POST':
        form = EditProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'อัพเดทวัสดุ สำเร็จ')
            return redirect('dashboard:products')
    else:
        form = EditProductForm(instance=product)
    context = {'title': 'Edit Product', 
               'form':form,
                'pending_orders_count': count_pending_orders(),}
    return render(request, 'edit_product.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def suppliers(request):
    suppliers = Suppliers.objects.all()

    query = request.GET.get('q')
    if query is not None:
        lookups = Q(supname__icontains=query) | Q(contactname__icontains=query)
        suppliers = Suppliers.objects.filter(lookups)


    page = request.GET.get('page')

    p = Paginator(suppliers, 20)
    try:
        suppliers = p.page(page)
    except:
        suppliers = p.page(1)

    context = {
        'title':'ซัพพลายเออร์', 
        'suppliers':suppliers,
        'pending_orders_count': count_pending_orders(),
        }
    return render(request, 'suppliers.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def add_suppliers(request):
    if request.method == 'POST':
        form = AddSuppliersForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'เพิ่มวัสดุเรียบร้อยแล้ว!')
            return redirect('dashboard:add_suppliers')
    else:
        form = AddSuppliersForm()
    context = {'title':'เพิ่มซัพพลายเอร์', 'form':form,
                'pending_orders_count': count_pending_orders(),}
    return render(request, 'add_suppliers.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def delete_suppliers(request, id):
    suppliers = Suppliers.objects.filter(id=id).delete()
    messages.success(request, 'ลบซัพพลายเออร์สำเร็จ')
    return redirect('dashboard:supplierss')



@user_passes_test(is_authorized_manager)
@login_required
def edit_suppliers(request, id):
    sup= Suppliers.objects.get(id=id)
    form = EditSuppliersForm()

    if request.method == 'POST':
        form = EditSuppliersForm(request.POST, request.FILES, instance=sup)
        if form.is_valid():
            form.save()
            messages.success(request, 'แก้ไขข้อมูล สำเร็จ')
            return redirect('dashboard:suppliers')
    else:
        form = EditSuppliersForm(instance=sup)
    context = {'title': 'แก้ไขข้อมูลซัพพลายเออร์', 
               'sup':sup,
               'form':form,
               'pending_orders_count': count_pending_orders(),}
    return render(request, 'edit_suppliers.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def detail_suppliers(request, id):
    sup = get_object_or_404(Suppliers, id=id)
    context = {
        'title':'รายละเอียดซัพพลายเออร์', 
        'sup':sup,
        'pending_orders_count': count_pending_orders(),
        }
    return render(request, 'detail_suppliers.html', context)



from openpyxl.utils import get_column_letter
from django.utils.encoding import escape_uri_path
@login_required
def export_receive_to_excel(request):
    now = datetime.now()
    month = int(request.GET.get('month', now.month))
    year_buddhist = int(request.GET.get('year', now.year + 543))
    year_ad = year_buddhist - 543

    queryset = Receiving.objects.filter(month=month, year=year_ad).order_by('-date_received')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายการรับวัสดุ"

    # ส่วนหัว
    headers = [
        'ลำดับ', 'รหัสวัสดุ', 'หมวดหมู่หลัก', 'ชื่อวัสดุ',
        'ข้อมูลซัพพลายเออร์',  # ✅ คอลัมน์รวม
        'จำนวนรับเข้า', 'จำนวนคงเหลือ', 'ราคา/หน่วย', 'หน่วย',
        'วันที่รับเข้า', 'วันที่อัปเดต', 'หมายเหตุ'
    ]
    ws.append(headers)

    # ข้อมูลแต่ละแถว
    for i, obj in enumerate(queryset, start=1):
        supplier_info = (
            f"{obj.suppliers.supname}\n"
            f"ชื่อผู้ติดต่อ: {obj.suppliers.contactname}\n"
            f"เบอร์โทร: {obj.suppliers.phone}\n"
            f"เลขภาษี: {obj.suppliers.taxnumber or '-'}\n"
            f"ที่อยู่: {obj.suppliers.address}"
        )

        ws.append([
            i,
            obj.product.product_id,
            obj.product.category.category.name_cate,
            obj.product.product_name,
            supplier_info,
            obj.quantityreceived,
            obj.quantity,
            float(obj.unitprice) if obj.unitprice else 0,
            obj.product.unit,
            obj.date_received.strftime('%d/%m/%Y %H:%M') if obj.date_received else '',
            obj.date_updated.strftime('%d/%m/%Y %H:%M'),
            obj.note or ''
        ])

    # ปรับความกว้างคอลัมน์
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 4

    # จัดรูปแบบข้อความให้ wrap เฉพาะคอลัมน์ "ข้อมูลซัพพลายเออร์" (คอลัมน์ที่ 5)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # สร้างชื่อไฟล์และ response
    filename = f"รายการรับเข้าวัสดุ_{month}_{year_buddhist}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'
    wb.save(response)
    return response


@user_passes_test(is_authorized_manager)
@login_required
def receive_list(request):
    now = datetime.now()

    # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    month = int(request.GET.get('month', now.month))
    year_buddhist = int(request.GET.get('year', now.year + 543))

    # แปลงปี พ.ศ. เป็น ค.ศ. สำหรับการค้นหาในฐานข้อมูล
    year_ad = year_buddhist - 543

    # ดึงข้อมูลรับเข้าสินค้าที่มีเดือนและปีที่ระบุสำหรับผู้ใช้งานปัจจุบัน
    # receive = Receiving.objects.all()
    
    # ดึงข้อมูลรับเข้าสินค้าที่มีเดือนและปีที่ระบุสำหรับผู้ใช้งานปัจจุบัน และเรียงจากวันที่รับเข้าใหม่สุดไปเก่าสุด
    receive = Receiving.objects.all().order_by('-date_received')

    # การค้นหา
    query = request.GET.get('q')
    if query is not None:
        lookups = Q(product__product_name__icontains=query) | Q(product__product_id__icontains=query)
        receive = receive.filter(lookups)
    else:
        receive = receive.filter(month=month, year=year_ad)

    context = {
        'title': 'รับเข้าวัสดุ',
        'receive': receive,
        'pending_orders_count': count_pending_orders(),
        'selected_month': month,
        'selected_year': year_buddhist,
        'years': range(2023 + 543, datetime.now().year + 1 + 543),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')],
        'month_name': thai_month_name(month),
    }
    previous_month = month - 1 if month > 1 else 12
    previous_year = year_ad if month > 1 else year_ad - 1
    context['previous_month_name'] = thai_month_name(previous_month)
    context['previous_year_buddhist'] = convert_to_buddhist_era(previous_year)

    return render(request, 'receive_list.html', context)




@user_passes_test(is_authorized_manager)
@login_required
def receive_product(request):
    if request.method == 'POST':
        form = ReceivingForm(request.POST, request.FILES)
        if form.is_valid():
            received_product = form.save(commit=False)
            received_product.save()
            
            # อัปเดตจำนวนสินค้าในสต็อก
            # product = received_product.product
            # stock, created = Stock.objects.get_or_create(product=product)
            # stock.quantity += received_product.quantityreceived
            # stock.save()

            # อัปเดตจำนวนทั้งหมดที่รับเข้า
            product = received_product.product
            total, created = Total_Quantity.objects.get_or_create(product=product)
            total.totalquantity += received_product.quantityreceived
            total.save()

            # อัปเดตจำนวนสินค้าในตาราง Product
            # Product.objects.filter(id=product.id).update(quantityinstock=F('quantityinstock') + received_product.quantity)

            messages.success(request, 'เพิ่มข้อมูลสำเร็จ')
            return redirect('dashboard:receive_list')  # หรือไปยังหน้าที่ต้องการ
        else:
            messages.error(request, 'เพิ่มข้อมูลไม่สำเร็จ')
    else:
        form = ReceivingForm()

    context = {
        'title':'รับเข้าวัสดุ',
        'form': form,
        'Product':Product.objects.all(),
        'Suppliers':Suppliers.objects.all(),
        'pending_orders_count': count_pending_orders(),
        }
    return render(request, 'receive_product.html', context)


@user_passes_test(is_authorized_manager)
@login_required
def update_received_product(request, id):
    received_product = get_object_or_404(Receiving, id=id)

    if request.method == 'POST':
        form = ReceivingForm(request.POST, request.FILES, instance=received_product)
        if form.is_valid():
            updated_received_product = form.save(commit=False)
            updated_received_product.save()

            # อัพเดทจำนวนสินค้าในสต็อก
            # product = updated_received_product.product
            # stock, created = Stock.objects.get_or_create(product=product)
            # stock.quantity -= received_product.quantityreceived  # ลบจำนวนเดิม
            # stock.quantity += updated_received_product.quantity  # เพิ่มจำนวนใหม่
            # stock.save()

            # อัพเดทจำนวนทั้งหมดที่รับเข้า
            # total, created = Total_Quantity.objects.get_or_create(product=product)
            # total.totalquantity -= received_product.quantityreceived  # ลบจำนวนเดิม
            # total.totalquantity += updated_received_product.quantityreceived  # เพิ่มจำนวนใหม่
            # total.save()

            # อัปเดตจำนวนสินค้าในตาราง Product
            # Product.objects.filter(id=product.id).update(quantityinstock=F('quantityinstock') + received_product.quantity)

            messages.success(request, 'อัพเดทข้อมูลสำเร็จ')
            return redirect('dashboard:receive_list')  # หรือไปยังหน้าที่ต้องการ
        else:
            messages.error(request, 'อัพเดทข้อมูลไม่สำเร็จ')
    else:
        form = ReceivingForm(instance=received_product)

    context = {
        'title':'อัพเดทการรับเข้าสินค้า',
        'form': form,
        'Product': Product.objects.all(),
        'Suppliers': Suppliers.objects.all(),
        'pending_orders_count': count_pending_orders(),
    }
    return render(request, 'update_received_product.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def delete_receive(request, id):
    suppliers = Receiving.objects.filter(id=id).delete()
    messages.success(request, 'ลบรับเข้าสำเร็จ')
    return redirect('dashboard:receive_list')



@user_passes_test(is_authorized_manager)
@login_required
def stock(request):
    products = Product.objects.all()
    
    # คำนวณ total_quantity สำหรับแต่ละสินค้า
    for product in products:
        product.total_quantity = Receiving.total_quantity_by_product(product.id)

    query = request.GET.get('q')
    if query is not None:
        lookups = Q(product_id__icontains=query) | Q(product_name__icontains=query)
        products = Product.objects.filter(lookups)

    page = request.GET.get('page')

    p = Paginator(products, 20)
    try:
        products = p.page(page)
    except:
        products = p.page(1)

    context = {
        'title':'สต๊อก', 
        'products':products,
        'pending_orders_count': count_pending_orders(),}
    return render(request, 'stock.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def workgroup(request):
    workgroup = WorkGroup.objects.all()

    query = request.GET.get('q')
    if query is not None:
        lookups = Q(work_group__icontains=query)
        workgroup = WorkGroup.objects.filter(lookups)

    page = request.GET.get('page')

    p = Paginator(workgroup, 20)
    try:
        workgroup = p.page(page)
    except:
        workgroup = p.page(1)

    context = {
        'title':'กลุ่มงาน', 
        'workgroup':workgroup,
        'pending_orders_count': count_pending_orders(),}
    return render(request, 'workgroup.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def add_work_group(request):
    if request.method == 'POST':
        form = WorkGroupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'เพิ่มหมวดหมู่เรียบร้อยแล้ว!')
            return redirect('dashboard:workgroup')  # เปลี่ยน 'success_page' เป็นชื่อ URL ของหน้าที่ต้องการไปหลังจากบันทึกข้อมูล
    else:
        form = WorkGroupForm()
    
    context = {'title':'เพิ่มกลุ่มงาน', 
               'form':form,
                'pending_orders_count': count_pending_orders(),}
    
    return render(request, 'add_work_group.html', context)


@user_passes_test(is_authorized_manager)
@login_required
def import_work_group(request):
    if request.method == 'POST':
        file = request.FILES['file']
        try:
            df = pd.read_excel(file)
            for index, row in df.iterrows():
                WorkGroup.objects.create(work_group=row['กลุ่มงาน'])
            
            messages.success(request, 'นำเข้าข้อมูลกลุ่มงานสำเร็จ')
        
        except Exception as e:
            messages.error(request, f'เกิดข้อผิดพลาด: {e}')
        
    return redirect('dashboard:workgroup')  



@user_passes_test(is_authorized_manager)
@login_required
def delete_workgroup(request, id):
    workgroup = WorkGroup.objects.filter(id=id).delete()
    messages.success(request, 'ลบกลุ่มงานสำเร็จ')
    return redirect('dashboard:workgroup')



@user_passes_test(is_authorized_manager)
@login_required
def edit_workgroup(request, id):
    work= WorkGroup.objects.get(id=id)
    form = EditWorkGroupForm()

    if request.method == 'POST':
        form = EditWorkGroupForm(request.POST, request.FILES, instance=work)
        if form.is_valid():
            form.save()
            messages.success(request, 'แก้ไขข้อมูลสำเร็จ')
            return redirect('dashboard:workgroup')
    else:
        form = EditWorkGroupForm(instance=work)
    context = {'title': 'แก้ไขข้อกลุ่มงาน', 
               'work':work,
               'form':form,
               'pending_orders_count': count_pending_orders(),}
    return render(request, 'edit_workgroup.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def category(request):
    category = Category.objects.all()

    query = request.GET.get('q')
    if query is not None:
        lookups = Q(name_cate__icontains=query)
        category = Category.objects.filter(lookups)

    page = request.GET.get('page')

    p = Paginator(category, 20)
    try:
        category = p.page(page)
    except:
        category = p.page(1)

    context = {
        'title':'หมวดหมู่หลัก', 
        'category':category,
        'pending_orders_count': count_pending_orders(),}
    return render(request, 'category.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def add_category(request):
    if request.method == 'POST':
        form = AddCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'เพิ่มหมวดหมู่เรียบร้อยแล้ว!')
            return redirect('dashboard:add_category')
    else:
        form = AddCategoryForm()
    context = {'title':'เพิ่มหมวดหมู่หลัก', 
               'form':form,
                'pending_orders_count': count_pending_orders(),}
    return render(request, 'add_category.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def delete_category(request, id):
    category = Category.objects.filter(id=id).delete()
    messages.success(request, 'ลบหมวดหมู่หลักสำเร็จ')
    return redirect('dashboard:category')



@user_passes_test(is_authorized_manager)
@login_required
def edit_category(request, id):
    cate= Category.objects.get(id=id)
    form = EditCategoryForm()

    if request.method == 'POST':
        form = EditCategoryForm(request.POST, request.FILES, instance=cate)
        if form.is_valid():
            form.save()
            messages.success(request, 'แก้ไขข้อมูล สำเร็จ')
            return redirect('dashboard:category')
    else:
        form = EditCategoryForm(instance=cate)
    context = {'title': 'แก้ไขข้อมูลหมวดหมู่หลัก', 
               'cate':cate,
               'form':form,
               'pending_orders_count': count_pending_orders(),}
    return render(request, 'edit_category.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def subcategory(request):
    subcategory = Subcategory.objects.all()

    query = request.GET.get('q')
    if query is not None:
        lookups = Q(name_sub__icontains=query)|Q(category__name_cate__icontains=query)
        subcategory = Subcategory.objects.filter(lookups)

    page = request.GET.get('page')

    p = Paginator(subcategory, 20)
    try:
        subcategory = p.page(page)
    except:
        subcategory = p.page(1)

    context = {
        'title':'หมวดหมู่ย่อย', 
        'subcategory':subcategory,
        'pending_orders_count': count_pending_orders(),}
    return render(request, 'subcategory.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def add_subcategory(request):
    if request.method == 'POST':
        form = AddSubcategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'เพิ่มหมวดหมู่ย่อยเรียบร้อยแล้ว!')
            return redirect('dashboard:add_subcategory')
    else:
        form = AddSubcategoryForm()
    context = {'title':'เพิ่มหมวดหมู่ย่อย', 
               'form':form,
                'pending_orders_count': count_pending_orders(),}
    return render(request, 'add_subcategory.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def delete_subcategory(request, id):
    category = Subcategory.objects.filter(id=id).delete()
    messages.success(request, 'ลบหมวดหมู่ย่อยสำเร็จ')
    return redirect('dashboard:subcategory')



@user_passes_test(is_authorized_manager)
@login_required
def edit_subcategory(request, id):
    sub= Subcategory.objects.get(id=id)
    form = EditSubcategoryForm()

    if request.method == 'POST':
        form = EditSubcategoryForm(request.POST, request.FILES, instance=sub)
        if form.is_valid():
            form.save()
            messages.success(request, 'แก้ไขข้อมูล สำเร็จ')
            return redirect('dashboard:subcategory')
    else:
        form = EditSubcategoryForm(instance=sub)
    context = {'title': 'แก้ไขข้อมูลหมวดหมู่ย่อย', 
               'sub':sub,
               'form':form,
               'pending_orders_count': count_pending_orders(),}
    return render(request, 'edit_subcategory.html', context)



@user_passes_test(is_authorized_manager)
@login_required
def orders_all(request):
    now = datetime.now()

#     # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
#     last_month = now.month if now.month > 1 else 12
#     last_year = now.year if now.month > 1 else now.year - 1

#     # ตรวจสอบว่ามีการระบุเดือนและปีในพารามิเตอร์ GET หรือไม่ ถ้าไม่มีใช้เดือนและปีของเดือนที่แล้ว
#     month = int(request.GET.get('month', last_month))
#     year_buddhist = int(request.GET.get('year', last_year + 543))

    # ใช้เดือนและปีปัจจุบันหากไม่ได้ระบุในพารามิเตอร์ GET
    month = int(request.GET.get('month', now.month))
    year_buddhist = int(request.GET.get('year', now.year + 543))

    # แปลงปี พ.ศ. เป็น ค.ศ. สำหรับการค้นหาในฐานข้อมูล
    year_ad = year_buddhist - 543

    # ดึงข้อมูลรับเข้าสินค้าที่มีเดือนและปีที่ระบุสำหรับผู้ใช้งานปัจจุบัน
    orders_all = Order.objects.all().select_related('user')

    # ตรวจสอบค่าจาก checkbox
    show_rejected = request.GET.get('show_rejected', None) == 'on'
    # หากไม่ต้องการแสดงรายการที่ถูกปฏิเสธให้กรองออก
    if not show_rejected:
        orders_all = orders_all.exclude(status=False)

    # การค้นหา
    query = request.GET.get('q')
    if query is not None:
        lookups = Q(id__icontains=query) | Q(user__first_name__icontains=query) | Q(user__last_name__icontains=query)
        orders_all = orders_all.filter(lookups)

    # กรองตามเดือนและปีที่ระบุ
    orders_all = orders_all.filter(
        month=month,
        year=year_ad
    )

    context = {
        'title': 'คำร้องเบิกวัสดุทั้งหมด',
        'orders_all': orders_all,
        'pending_orders_count': count_pending_orders(),
        'selected_month': month,
        'selected_year': year_buddhist,
        'years': range(2023 + 543, datetime.now().year + 1 + 543),
        'months': [
            (1, 'มกราคม'), (2, 'กุมภาพันธ์'), (3, 'มีนาคม'), (4, 'เมษายน'),
            (5, 'พฤษภาคม'), (6, 'มิถุนายน'), (7, 'กรกฎาคม'), (8, 'สิงหาคม'),
            (9, 'กันยายน'), (10, 'ตุลาคม'), (11, 'พฤศจิกายน'), (12, 'ธันวาคม')],
        'month_name': thai_month_name(month),
        'show_rejected': show_rejected,  # เพิ่มค่า show_rejected ใน context
    }

    previous_month = month - 1 if month > 1 else 12
    previous_year = year_ad if month > 1 else year_ad - 1
    context['previous_month_name'] = thai_month_name(previous_month)
    context['previous_year_buddhist'] = convert_to_buddhist_era(previous_year)

    return render(request, 'orders_all.html', context)




@user_passes_test(is_authorized_manager)
@login_required
def orders(request):
    orders = Order.objects.all()

    query = request.GET.get('q')
    if query is not None:
        lookups = Q(id__icontains=query)|Q(date_created__icontains=query)
        orders = Order.objects.filter(lookups).order_by('id')

    page = request.GET.get('page')
    p = Paginator(orders, 20)
    try:
        orders = p.page(page)
    except:
        orders = p.page(1)

    # ดึงข้อมูลออร์เดอร์ทั้งหมดที่รอการยืนยัน
    orders_waiting_status = Order.objects.filter(status=None).order_by('id')

    # ส่งข้อมูล approval_count ไปยัง template
    context = {
        'title':'คำร้องใหม่ รออนุมัติ..',
        'orders':'orders',
        'orders':orders_waiting_status,
        'status_count': orders_waiting_status.count(),
        'pending_orders_count': count_pending_orders(),
        }
    return render(request, 'orders.html', context)



# ใหม่4
# orders/views.py
@user_passes_test(is_authorized_manager)
@login_required
def approve_orders(request, order_id):
    ap = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        form = ApproveForm(request.POST, instance=ap)
        if form.is_valid():
            old_status = ap.status  # เก็บสถานะเดิมก่อนการเปลี่ยนแปลง
            new_status = form.cleaned_data.get('status')

            if new_status == False:
                print("คำสั่งซื้อถูกปฏิเสธ กำลังกู้คืนสต็อก......")
                # คืนจำนวนสินค้ากลับไปยังสต๊อก
                for item in ap.items.all():
                    # product = item.product
                    # product.quantityinstock += item.quantity
                    # product.save()
                    # print(f"Restored {item.quantity} of {product.product_name} to stock.")

                    # คืนจำนวนสินค้าที่รับเข้าใน Receiving
                    receiving = item.receiving
                    receiving.quantity += item.quantity
                    receiving.save()
                    print(f"กู้คืน {item.quantity} to receiving ID {receiving.id}.")
            form.save()
            messages.success(request, 'ดำเนินการสำเร็จ')

            # Notify the user about the approval
            notify_user_approved(ap.id)
            notify_admin_order_status(ap.id)
            
            return redirect(reverse('dashboard:order_detail', args=[order_id]))
        else:
            messages.error(request, 'ดำเนินการไม่สำเร็จ')
    else:
        form = ApproveForm(instance=ap)
        
    return render(request, 'orders.html', {
        'ap': ap,
        'form': form,
        'title': 'แก้ไขข้อมูลสมาชิก',
        'pending_orders_count': count_pending_orders(),
    })


@login_required
def approve_pay(request, order_id):
    ap = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        form = ApprovePayForm(request.POST, instance=ap)
        if form.is_valid():
            print("จ่ายวัสดุแล้ว")
            form.save()
            notify_user_pay_confirmed(order_id)  # เรียกฟังก์ชันการแจ้งเตือน
            # messages.success(request, 'ยืนยันการจ่ายวัสดุสำเร็จ')
            return redirect(reverse('dashboard:orders_all') + '?success=true')
        else:
            messages.error(request, 'ดำเนินการไม่สำเร็จ')
    else:
        form = ApprovePayForm(instance=ap)
        
    return render(request, 'orders_all.html', {
        'ap': ap,
        'form': form,
    })


@user_passes_test(is_authorized_manager)
@login_required
def order_detail(request, id):
    order = Order.objects.filter(id=id).first()
    items = Issuing.objects.filter(order=order).all()
    context = {
        'title':'รายการเบิกจ่ายวัสดุ', 
        'items':items, 
        'order':order,
        'pending_orders_count': count_pending_orders(),
        }
    return render(request, 'order_detail.html', context)


from django.contrib.admin.views.decorators import staff_member_required


@login_required
def notification(request):
    # ดึงข้อมูลวัสดุที่หมด (จำนวนคงเหลือ = 0)
    products = Product.objects.filter(quantityinstock=0)

    # ดึงข้อมูลทั้งหมดพร้อมข้อมูลผู้ใช้งานและรายการวัสดุ โดยเรียงลำดับจากคำร้องที่ยื่นมาก่อน
    # notification = OutOfStockNotification.objects.select_related('user', 'product').order_by('date_created')  # เรียงตามเวลาที่สร้าง (จากเก่าไปใหม่)
    notification = OutOfStockNotification.objects.select_related('user', 'product')

    # pagination
    page = request.GET.get('page')
    p = Paginator(notification, 15)
    try:
        notification = p.page(page)
    except:
        notification = p.page(1)

    context = {
        'title': 'รายการแจ้งวัสดุหมด', 
        'notification': notification,  # เปลี่ยนเป็น notification
        'product': products,  # ส่งข้อมูลวัสดุไปยัง Template
    }
    return render(request, 'notification.html', context)



from django.views.decorators.csrf import csrf_exempt

# รับทราบการแจ้งเตือนสำหรับวัสดุ
@csrf_exempt
@login_required
@staff_member_required
def acknowledge_notification(request, notification_id):
    notification = get_object_or_404(OutOfStockNotification, id=notification_id)

    if request.method == "POST":
        acknowledged_note = request.POST.get("acknowledged_note", "").strip()
        if not notification.acknowledged:
            notification.acknowledged = True
            notification.acknowledged_note = acknowledged_note
            notification.save()

         # เรียกใช้ฟังก์ชันแจ้งเตือน
            send_acknowledge_notification(notification)

            messages.success(request, f"รับทราบการแจ้งเตือนสำหรับวัสดุ '{notification.product.product_name}' เรียบร้อยแล้ว!")
        else:
            messages.warning(request, f"การแจ้งเตือนนี้ได้รับการรับทราบแล้วก่อนหน้านี้!")

    return redirect('dashboard:notification')


# ยืนยันการเติมสต๊อกวัสดุ
@csrf_exempt
@login_required
@staff_member_required
def restock_notification(request, notification_id):
    notification = get_object_or_404(OutOfStockNotification, id=notification_id)

    if request.method == "POST":
        acknowledged_note = request.POST.get("acknowledged_note", "").strip()
        if not notification.restocked:
            notification.restocked = True
            notification.acknowledged_note = acknowledged_note
            notification.save()

            # เรียกใช้ฟังก์ชันแจ้งเตือน
            send_restock_notification(notification)
            
            messages.success(request, f"ยืนยันการเติมสต๊อกวัสดุ '{notification.product.product_name}' สำเร็จแล้ว!")
        else:
            messages.warning(request, f"การเติมสต๊อกสำหรับวัสดุ '{notification.product.product_name}' ได้ดำเนินการแล้วก่อนหน้านี้!")

    return redirect('dashboard:notification')